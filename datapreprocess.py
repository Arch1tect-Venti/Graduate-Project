import pandas as pd
import xlrd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import xlwt  # 进行excel操作
import tensorflow as tf
# 学生企业二维关系数据转化为一维
def data2dto1d():
    # 读取文件
    wb = xlrd.open_workbook(filename="data/cominfo_1.xlsx")
    # 读取对应sheet
    CLsheet = wb.sheet_by_name("二维数据")


# 企业属性的独热编码
def One_hot_code_comp():
    # 读取企业信息的文件
    wb = xlrd.open_workbook(filename="data/cominfo.xlsx")
    # 读取对应sheet
    CLsheet = wb.sheet_by_name("Sheet1")
    # # 读取行数
    # max_rownum=CLsheet.nrows
    # print(max_rownum)
    df = pd.DataFrame(columns=['公司名称','法人','注册资本','地址'])
    df['公司名称'] = CLsheet.col_values(0)
    df['法人'] = CLsheet.col_values(1)
    df['注册资本'] = CLsheet.col_values(2)
    df['成立日期'] = CLsheet.col_values(3)
    df['地址'] = CLsheet.col_values(4)

    # x = CLsheet.cell(0, 0).value

    dummies1 = pd.get_dummies(df['公司名称'],prefix='key')
    dummies2 = pd.get_dummies(df['法人'],prefix='key')
    dummies3 = pd.get_dummies(df['注册资本'],prefix='key')
    dummies4 = pd.get_dummies(df['成立日期'],prefix='key')
    dummies5 = pd.get_dummies(df['地址'],prefix='key')

    # 拼接
    df_save = pd.concat([dummies1, dummies2, dummies3,dummies4,dummies5], axis=1)
    # 存储独热向量
    wb_save = Workbook()
    ws_save = wb_save.active
    ws_save.title = '独热编码_公司'
    for i in dataframe_to_rows(df_save,index=True,header=True):
        ws_save.append(i)
    wb_save.save("data/One_hot_data_comp.xlsx")

# 学生属性的独热编码
def One_hot_code_std():
    # 读取学生信息的文件
    wb = xlrd.open_workbook(filename="data/stdinfo.xlsx")
    # 读取对应sheet
    CLsheet = wb.sheet_by_name("Sheet1")
    # # 读取行数
    # max_rownum=CLsheet.nrows
    # print(max_rownum)

    # 这里应该加一个去表头的

    df = pd.DataFrame(columns=['姓名','学院','专业','性别','民族','定向生','DXDWMC','DXDWDM','生源地址','生源地市','方向'])
    df['姓名'] = CLsheet.col_values(0)
    df['学院'] = CLsheet.col_values(1)
    df['专业'] = CLsheet.col_values(2)
    df['性别'] = CLsheet.col_values(3)
    df['民族'] = CLsheet.col_values(4)
    df['定向生'] = CLsheet.col_values(5)
    df['DXDWMC'] = CLsheet.col_values(6)
    df['DXDWDM'] = CLsheet.col_values(7)
    df['生源地址'] = CLsheet.col_values(8)
    df['生源地市'] = CLsheet.col_values(9)
    df['方向'] = CLsheet.col_values(10)

    dummies1 = pd.get_dummies(df['姓名'],prefix='key')
    dummies2 = pd.get_dummies(df['学院'],prefix='key')
    dummies3 = pd.get_dummies(df['专业'],prefix='key')
    dummies4 = pd.get_dummies(df['性别'],prefix='key')
    dummies5 = pd.get_dummies(df['民族'],prefix='key')
    dummies6 = pd.get_dummies(df['定向生'],prefix='key')
    dummies7 = pd.get_dummies(df['DXDWMC'],prefix='key')
    dummies8 = pd.get_dummies(df['DXDWDM'],prefix='key')
    dummies9 = pd.get_dummies(df['生源地址'],prefix='key')
    dummies10 = pd.get_dummies(df['生源地市'],prefix='key')
    dummies10 = pd.get_dummies(df['方向'],prefix='key')

    # 拼接
    df_save = pd.concat([dummies1, dummies2, dummies3,dummies4,dummies5,dummies6,dummies7,dummies8,dummies9,dummies10], axis=1)
    # 存储独热向量
    wb_save = Workbook()
    ws_save = wb_save.active
    ws_save.title = '独热编码_学生'
    for i in dataframe_to_rows(df_save,index=True,header=True):
        ws_save.append(i)
    wb_save.save("data/One_hot_data_std.xlsx")

if __name__ == "__main__":
    # One_hot_code_comp()
    # One_hot_code_std()
    data2dto1d()
    print("完毕！")